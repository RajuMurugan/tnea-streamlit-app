import streamlit as st  
import pandas as pd
import os
import io
import plotly.express as px
from openpyxl import load_workbook

st.set_page_config(page_title="TNEA Vacancy Viewer", layout="wide")
st.title("🎓 TNEA Vacancy List")

# --- Load Excel file with all sheets ---
excel_file = r"E:\Raju_M\Python\TNEA_2024\PDF_Files\Vacancy_list\all_vacancies.xlsx"

def load_excel_sheets_safe(path):
    wb = load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    data_dict = {}

    for sheet in sheet_names:
        ws = wb[sheet]
        data = list(ws.values)
        if not data:
            continue  # skip empty sheets
        header = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=header)
        data_dict[sheet] = df

    return sheet_names, data_dict

sheet_names, data_dict = load_excel_sheets_safe(excel_file)

# ----------------------------- CATEGORY 1 -----------------------------
st.markdown("## 🗂️ Select Branch and Community")

col_cat1_1, col_cat1_2, col_cat1_3 = st.columns(3)

with col_cat1_1:
    selected_sheet_1 = st.selectbox("📂 Select Vacancy - Category", sheet_names, key="cat1_sheet")
    df1 = data_dict[selected_sheet_1]

with col_cat1_2:
    pass  # UI spacing

if df1.empty:
    st.error("❌ No data found in the selected sheet.")
    st.stop()

# --- Clean column names ---
df1.columns = [str(col).strip().upper().replace("  ", " ").replace("\n", " ") for col in df1.columns]

rename_map = {}
for col in df1.columns:
    if "COLLEGE CODE" in col:
        rename_map[col] = 'College Code'
    elif "COLLEGE NAME" in col:
        rename_map[col] = 'College Name'
    elif "BRANCH CODE" in col:
        rename_map[col] = 'Branch Code'
    elif "BRANCH NAME" in col:
        rename_map[col] = 'Branch Name'

df1.rename(columns=rename_map, inplace=True)

required_id_vars = ['College Name', 'College Code', 'Branch Code', 'Branch Name']
community_cols = ['OC', 'BC', 'BCM', 'MBC', 'SC', 'SCA', 'ST']
df1 = df1[[col for col in df1.columns if col in required_id_vars + community_cols]]
df1_melted = df1.melt(id_vars=required_id_vars, value_vars=community_cols, var_name='Community', value_name='Seats')
df1_melted['Seats'] = pd.to_numeric(df1_melted['Seats'], errors='coerce').fillna(0).astype(int)

with col_cat1_2:
    branch_codes = sorted(df1_melted['Branch Code'].dropna().unique())
    selected_branch_1 = st.selectbox("🔍 Select Branch Code", branch_codes)

with col_cat1_3:
    community_options = ['All'] + sorted(df1_melted['Community'].dropna().unique())
    selected_community_1 = st.selectbox("🧑‍🤝‍🧑 Filter by Community (Optional)", community_options)

# --- Filtered data for Category 1 ---
branch_df = df1_melted[df1_melted['Branch Code'] == selected_branch_1]
if selected_community_1 != "All":
    branch_df = branch_df[branch_df['Community'] == selected_community_1]

if not branch_df.empty:
    branch_name = branch_df['Branch Name'].iloc[0]
    st.header(f"📘 Summary for Branch: {selected_branch_1} - {branch_name}")

    summary_df = branch_df.groupby('Community')['Seats'].sum().reset_index().sort_values(by='Seats', ascending=False)
    total_seats = summary_df['Seats'].sum()
    summary_df.loc[len(summary_df.index)] = ['Total', total_seats]

    fig = px.bar(summary_df[summary_df['Community'] != 'Total'], x='Community', y='Seats', color='Community',
                 title=f"Community-wise Seat Distribution (Total: {total_seats} seats)",
                 labels={'Seats': 'Number of Seats'}, height=400)
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("🧾 College-wise Seat Data")
    st.dataframe(branch_df, use_container_width=True)

    excel_buffer = io.BytesIO()
    branch_df.to_excel(excel_buffer, index=False, engine='openpyxl')
    excel_buffer.seek(0)

    st.download_button(
        label="📥 Download Branch Summary",
        data=excel_buffer,
        file_name=f"{selected_branch_1}_Community_Seats.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("⚠️ No data found for the selected branch/community.")

# ----------------------------- CATEGORY 2 -----------------------------
st.markdown("---")
st.markdown("## 🏫 Select College and Branch")

col_cat2_1, col_cat2_2, col_cat2_3 = st.columns(3)

with col_cat2_1:
    selected_sheet_2 = st.selectbox("📂 Select Vacancy - Category", sheet_names, key="cat2_sheet")
    df2 = data_dict[selected_sheet_2]

if df2.empty:
    st.error("❌ No data found in the selected sheet.")
    st.stop()

df2.columns = [str(col).strip().upper().replace("  ", " ").replace("\n", " ") for col in df2.columns]
df2.rename(columns=rename_map, inplace=True)

df2 = df2[[col for col in df2.columns if col in required_id_vars + community_cols]]
df2_melted = df2.melt(id_vars=required_id_vars, value_vars=community_cols, var_name='Community', value_name='Seats')
df2_melted['Seats'] = pd.to_numeric(df2_melted['Seats'], errors='coerce').fillna(0).astype(int)

# Combine College Code - Name
df2_melted['College Combined'] = df2_melted['College Code'].astype(str) + ' - ' + df2_melted['College Name']
unique_colleges = sorted(df2_melted['College Combined'].dropna().unique())

with col_cat2_2:
    selected_college_combined = st.selectbox("🏫 Select College (Code - Name)", ['All'] + unique_colleges)

with col_cat2_3:
    branch_codes_2 = sorted(df2_melted['Branch Code'].dropna().unique())
    selected_branch_code = st.selectbox("🔍 Filter by Branch Code (Optional)", ['All'] + branch_codes_2)

selected_code, selected_name = "All", "All"
college_df = df2_melted.copy()

if selected_college_combined != "All":
    selected_code, selected_name = selected_college_combined.split(" - ", 1)
    college_df = college_df[
        (college_df['College Code'].astype(str) == selected_code.strip()) &
        (college_df['College Name'].str.strip() == selected_name.strip())
    ]

if selected_branch_code != "All":
    college_df = college_df[college_df['Branch Code'] == selected_branch_code]

if not college_df.empty:
    st.subheader("🏫 College-wise Community Seat Distribution")
    st.dataframe(college_df, use_container_width=True)

    summary2 = college_df.groupby('Community')['Seats'].sum().reset_index()
    total2 = summary2['Seats'].sum()
    summary2.loc[len(summary2.index)] = ['Total', total2]

    fig2 = px.bar(summary2[summary2['Community'] != 'Total'], x='Community', y='Seats', color='Community',
                  title=f"Community-wise Seat Distribution for College (Total: {total2} seats)",
                  labels={'Seats': 'Number of Seats'}, height=400)
    st.plotly_chart(fig2, use_container_width=True)

    excel_buffer2 = io.BytesIO()
    college_df.to_excel(excel_buffer2, index=False, engine='openpyxl')
    excel_buffer2.seek(0)

    st.download_button(
        label="📥 Download College Summary",
        data=excel_buffer2,
        file_name=f"{selected_code.strip()}_{selected_name.strip().replace(' ', '_')}_Seats.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("⚠️ No data found for the selected college or branch.")
