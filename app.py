import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import yaml
import requests
import io
import uuid
import time
from datetime import timedelta, datetime
import os
import plotly.express as px
from openpyxl import load_workbook
import json
from zoneinfo import ZoneInfo

# -------------------------------------------------
# ✅ Page Config (MUST BE FIRST Streamlit command)
# -------------------------------------------------
st.set_page_config(page_title="TNEA Full App", layout="wide")

# -------------------------------------------------
# ✅ PREMIUM FLAG (from URL)
# -------------------------------------------------
premium_flag = "0"
try:
    premium_flag = st.query_params.get("premium", "0")  # new streamlit
except:
    premium_flag = st.experimental_get_query_params().get("premium", ["0"])[0]  # old streamlit

is_premium = str(premium_flag) == "1"

# -------------------------------------------------
# ✅ MODE DISPLAY + BUTTON (same row)
# -------------------------------------------------
col_mode, col_btn = st.columns([3, 1])

with col_mode:
    if is_premium:
        st.success("✅ PREMIUM MODE ENABLED")
    else:
        st.info("🆓 NORMAL MODE (Free User)")

with col_btn:
    st.markdown("<br>", unsafe_allow_html=True)
    if not is_premium:
        st.link_button("💳 Go Premium", "https://tnea-choice-list.streamlit.app/?premium=1")
        st.markdown("💳 Lifetime Premium: **₹299 (One Time Payment)**")

# -------------------------------------------------
# ✅ Style Settings
# -------------------------------------------------
st.markdown("""
    <style>
    @media (max-width: 768px) {
        .nav-link span { display: inline !important; }
    }
    .stDataFrame div { color: black !important; }
    </style>
""", unsafe_allow_html=True)

# -------------------------------------------------
# ✅ File Paths
# -------------------------------------------------
base_path = "./"
config_path = base_path + "config.yaml"
device_session_path = base_path + "device_session.yaml"
chat_path = base_path + "chat_messages.json"

SESSION_TIMEOUT = 180  # 3 minutes

# -------------------------------------------------
# ✅ Load Config.yaml (Login optional)
# -------------------------------------------------
try:
    with open(config_path) as file:
        config = yaml.safe_load(file)
    user_data = config["credentials"]["users"]
except Exception:
    user_data = {}
    st.warning("⚠️ config.yaml not loaded (Login will not work)")

# -------------------------------------------------
# ✅ Load device_session.yaml
# -------------------------------------------------
try:
    with open(device_session_path) as session_file:
        session_data = yaml.safe_load(session_file)
except Exception:
    session_data = {"active_users": {}}

# -------------------------------------------------
# ✅ Load chat_messages.json
# -------------------------------------------------
if not os.path.exists(chat_path):
    with open(chat_path, "w") as f:
        json.dump([], f)

# -------------------------------------------------
# ✅ Session init
# -------------------------------------------------
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "mobile" not in st.session_state:
    st.session_state.mobile = "Guest"
if "device_id" not in st.session_state:
    st.session_state.device_id = str(uuid.uuid4())

# =================================================
# ✅ MENU
# =================================================

st.markdown("""
    <h2 style='text-align: center; color: #0d6efd; font-weight: bold;'>
        🔽 Select a Feature Below 🔽
    </h2>
""", unsafe_allow_html=True)

col1, col2, col3 = st.columns([1, 2, 1])

with col2:
    # ✅ Menu options based on Normal / Premium mode
    if is_premium:
        menu_options = [
            "Home",
            "Cutoff Calculator",
            "Branch List",
            "TNEA College List (PDF)",
            "Create TNEA Choice List",
            "TNEA Vacancy Seat Matrix"
        ]
        menu_icons = [
            "house",
            "calculator",
            "list-check",
            "building",
            "list-check",
            "table"
        ]
    else:
        menu_options = [
            "Home",
            "Cutoff Calculator",
            "Branch List",
            "TNEA College List (PDF)"
        ]
        menu_icons = [
            "house",
            "calculator",
            "list-check",
            "building"
        ]

    selected = option_menu(
        menu_title=None,
        options=menu_options,
        icons=menu_icons,
        default_index=0,
        orientation="horizontal"
    )

# =================================================
# ✅ PAGE 1: HOME
# =================================================
if selected == "Home":

    # ✅ Welcome Text (ONLY ONE TIME)
    st.markdown("""
        <h1 style='text-align: center; font-weight: bold;'>📘 Welcome to TNEA SmartGuide 2026 </h1>

        <h2 style="text-align:center; color:#0d6efd;">📌 About TNEA – Tamil Nadu Engineering Admissions</h2>

        <div style="font-size:17px; line-height:1.8; text-align:justify; padding:10px;">

        <b>TNEA (Tamil Nadu Engineering Admissions)</b> is the official counselling process for admission into 
        <b>B.E / B.Tech</b> courses in Tamil Nadu Government, Government Aided, and Self-Financing engineering colleges.

        <br><br>

        ✅ Admission is mainly based on <b>Class 12 Marks</b> (Maths, Physics, Chemistry) and the <b>Cutoff Score (out of 200)</b>.  
        <br><br>

        📌 Through TNEA SmartGuide 2026 , students can:
        <ul>
        <li>Check previous year cutoff trends</li>
        <li>Compare colleges and departments</li>
        <li>Create a better <b>choice list</b> for counselling</li>
        <li>Analyze <b>vacancy seat matrix</b> round-wise</li>
        </ul>
        <br>
        🎯 <b>This app is made to help Tamil Nadu students for TNEA 2026 admissions</b> by providing cutoff tools,
        choice list support, and counselling insights in one place.
        </div>
        <h2 style="text-align:center; color:#0d6efd;">📌 TNEA – தமிழ்நாடு பொறியியல் சேர்க்கை</h2>

        <div style="font-size:17px; line-height:1.8; text-align:justify; padding:10px;">

        <b>TNEA (Tamil Nadu Engineering Admissions)</b> என்பது தமிழ்நாட்டில் உள்ள 
        <b>B.E / B.Tech</b> படிப்புகளுக்கான அதிகாரப்பூர்வ கலந்தாய்வு (Counselling) முறையாகும்.

        <br><br>

        ✅ மாணவர்களின் <b>12ஆம் வகுப்பு மதிப்பெண்கள்</b> மற்றும் <b>Cutoff மதிப்பெண் (200ல்)</b> அடிப்படையில் சேர்க்கை வழங்கப்படுகிறது.

        <br><br>

        📌 இந்த TNEA SmartGuide 2026 மூலம் நீங்கள்:
        <ul>
        <li>Cutoff கணக்கிடலாம்</li>
        <li>கல்லூரி / Branch cutoff பார்க்கலாம்</li>
        <li>Choice List உருவாக்கலாம்</li>
        <li>Vacancy Seat Matrix ஆய்வு செய்யலாம்</li>
        </ul>

        </div>

        <div style='text-align: center; font-size: 18px; margin-top: 20px;'>
            <b>✅ Create TNEA Choice List</b> – Filter colleges by cutoff, department, and community<br><br>
            <b>📊 TNEA Vacancy Seat Matrix</b> – Analyze vacant seats by branch, college, and community<br><br>
            📞 Contact: +91-8248696926<br>
            📧 Email: rajumurugannp@gmail.com<br>
            👨‍💻 Developed by Dr. Raju Murugan<br><br>
            &copy; 2026 TNEA Info App. All rights reserved.
        </div>
    """, unsafe_allow_html=True)

    # ✅ Premium Offer (ONLY Free)
    if not is_premium:
        st.markdown("---")
        st.markdown("## 🔒 Premium Features")
        st.warning("Premium unlocks: ✅ Choice List + ✅ Vacancy Seat Matrix")
        st.markdown("💳 Lifetime Premium: **₹299 (One Time Payment)**")
        st.info("✅ Unlock Premium Features by clicking below 👇")

        st.link_button("💳 Go Premium", "https://tnea-choice-list.streamlit.app/?premium=1")


    # ✅ Previous Year Question Papers (FREE for all)
    st.markdown("---")
    st.markdown("### 📚 Very useful web Links")

    st.markdown("""
    <div style='background-color: #f9f9f9; padding: 15px; border-left: 8px solid #4CAF50; border-radius: 10px; font-size: 16px;'>
    📘 <a href='https://globaleduhub4u.blogspot.com/2025/03/anna-university-previous-year-questions.html' target='_blank'
        style='text-decoration: none; color: #007bff; font-weight: bold;'>Anna University Previous Year Question Papers</a><br>
    📗 <a href='https://globaleduhub4u.blogspot.com/p/gate-previous-year-qps.html' target='_blank'
        style='text-decoration: none; color: #007bff; font-weight: bold;'>GATE Previous Year Question Papers</a><br>
    📘 <a href='https://globaleduhub4u.blogspot.com/2025/03/numberiq.html' target='_blank'
        style='text-decoration: none; color: #007bff; font-weight: bold;'>Check Your Maths IQ</a><br>
    📗 <a href='https://static.tneaonline.org/docs/7_List_of_TFCs.pdf?t=1768660504718' target='_blank'
        style='text-decoration: none; color: #007bff; font-weight: bold;'>LIST OF TNEA FACILITATION CENTERS</a><br>
    📘 <a href='https://static.tneaonline.org/docs/TNEA_Tent_Schedule_2025.pdf' target='_blank'
        style='text-decoration: none; color: #007bff; font-weight: bold;'>TNEA 2025 Schedule </a><br>
    </div>
    """, unsafe_allow_html=True)

    # ✅ Chat Feature
    st.markdown("---")
    st.subheader("💬 Community Chat Room")

    with open(chat_path, "r") as f:
        chat_data = json.load(f)

    for entry in chat_data[-100:]:
        st.markdown(f"**{entry.get('user','Guest')}**: {entry.get('message','')}")

    new_message = st.text_input("Type your message...", key="chat_message_input")

    if st.button("Send", key="chat_send_btn") and new_message.strip():
        chat_data.append({"user": st.session_state.mobile, "message": new_message.strip()})
        with open(chat_path, "w") as f:
            json.dump(chat_data, f, indent=2)
        st.rerun()

# =================================================
# ✅ PAGE 2: CUTOFF CALCULATOR (FREE + PREMIUM)
# =================================================
elif selected == "Cutoff Calculator":
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    st.markdown("## 📚 TNEA 2026 Cut off Mark Calculator")

    ist_time_str = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y %I:%M %p")

    def reset_marks():
        st.session_state.pop("student_name", None)
        st.session_state.pop("maths_mark", None)
        st.session_state.pop("physics_mark", None)
        st.session_state.pop("chemistry_mark", None)

    st.markdown("### ✍️ Enter Student Details")

    name = st.text_input("👤 Student Name", placeholder="Enter your name", key="student_name")

    col1, col2, col3 = st.columns(3)
    with col1:
        maths = st.text_input("✏️ Maths", placeholder="Enter Maths mark", key="maths_mark")
    with col2:
        physics = st.text_input("⚡ Physics", placeholder="Enter Physics mark", key="physics_mark")
    with col3:
        chemistry = st.text_input("🧪 Chemistry", placeholder="Enter Chemistry mark", key="chemistry_mark")

    b1, b2 = st.columns([1, 1])
    with b1:
        calc_btn = st.button("✅ Calculate Cutoff", key="calc_cutoff_btn")
    with b2:
        st.button("🔄 Reset", on_click=reset_marks, key="reset_cutoff_btn")

    def generate_pdf(student_name, maths_val, physics_val, chemistry_val, cutoff_val):
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, height - 70, "TNEA Cutoff Mark Result (2026)")

        c.setFont("Helvetica", 12)
        c.drawString(50, height - 110, f"Name: {student_name}")
        c.drawString(50, height - 130, f"Date & Time (IST): {ist_time_str}")

        c.line(50, height - 150, 550, height - 150)

        c.setFont("Helvetica-Bold", 13)
        c.drawString(50, height - 190, "Marks Details:")

        c.setFont("Helvetica", 12)
        c.drawString(50, height - 220, f"Maths: {maths_val:.2f} / 100")
        c.drawString(50, height - 245, f"Physics: {physics_val:.2f} / 100  →  {physics_val/2:.2f} / 50")
        c.drawString(50, height - 270, f"Chemistry: {chemistry_val:.2f} / 100  →  {chemistry_val/2:.2f} / 50")

        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, height - 320, f"Final Cutoff: {cutoff_val:.2f} / 200")

        c.showPage()
        c.save()
        buffer.seek(0)
        return buffer

    def generate_image(student_name, maths_val, physics_val, chemistry_val, cutoff_val):
        img = Image.new("RGB", (900, 520), "white")
        draw = ImageDraw.Draw(img)

        try:
            font_title = ImageFont.truetype("arial.ttf", 36)
            font_text = ImageFont.truetype("arial.ttf", 22)
        except:
            font_title = ImageFont.load_default()
            font_text = ImageFont.load_default()

        draw.text((30, 20), "TNEA Cutoff Mark Result (2026)", font=font_title, fill="black")
        draw.text((30, 90), f"Name: {student_name}", font=font_text, fill="black")
        draw.text((30, 120), f"Date & Time (IST): {ist_time_str}", font=font_text, fill="black")

        draw.line((30, 160, 870, 160), fill="black", width=2)

        draw.text((30, 190), f"Maths: {maths_val:.2f} / 100", font=font_text, fill="black")
        draw.text((30, 230), f"Physics: {physics_val:.2f} / 100  →  {physics_val/2:.2f} / 50", font=font_text, fill="black")
        draw.text((30, 270), f"Chemistry: {chemistry_val:.2f} / 100  →  {chemistry_val/2:.2f} / 50", font=font_text, fill="black")

        draw.line((30, 320, 870, 320), fill="black", width=2)
        draw.text((30, 350), f"Final Cutoff: {cutoff_val:.2f} / 200", font=font_title, fill="black")

        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer

    if calc_btn:
        try:
            if name.strip() == "":
                st.error("❌ Please enter Student Name.")
            else:
                maths_val = float(maths)
                physics_val = float(physics)
                chemistry_val = float(chemistry)

                if not (0 <= maths_val <= 100 and 0 <= physics_val <= 100 and 0 <= chemistry_val <= 100):
                    st.error("❌ Please enter marks between 0 and 100 only.")
                else:
                    cutoff_val = maths_val + (physics_val / 2) + (chemistry_val / 2)

                    st.success(f"🎯 {name} - Your TNEA Cutoff Mark is: **{cutoff_val:.2f} / 200**")
                    st.caption(f"🕒 Generated Time (IST): {ist_time_str}")

                    st.markdown("### 📥 Download Result")

                    pdf_file = generate_pdf(name, maths_val, physics_val, chemistry_val, cutoff_val)
                    img_file = generate_image(name, maths_val, physics_val, chemistry_val, cutoff_val)

                    d1, d2 = st.columns(2)
                    with d1:
                        st.download_button(
                            "⬇️ Download PDF",
                            data=pdf_file,
                            file_name=f"TNEA_Cutoff_2026_{name.replace(' ', '_')}.pdf",
                            mime="application/pdf"
                        )
                    with d2:
                        st.download_button(
                            "⬇️ Download Image (PNG)",
                            data=img_file,
                            file_name=f"TNEA_Cutoff_2026_{name.replace(' ', '_')}.png",
                            mime="image/png"
                        )
        except:
            st.error("❌ Please enter valid numbers in Maths / Physics / Chemistry.")

# =================================================
# ✅ PAGE 3: BRANCH LIST (FREE)
# =================================================
elif selected == "Branch List":

    st.markdown("## 📌 Branch Code & Branch Name List (TNEA)")
    st.caption("Search Branch Code or Branch Name ✅")
    branch_data = [ {"Branch Code": "AO", "Branch Name": "AEROSPACE ENGINEERING"}, {"Branch Code": "AE", "Branch Name": "AERONAUTICAL ENGINEERING"}, {"Branch Code": "AG", "Branch Name": "AGRICULTURAL ENGINEERING"}, {"Branch Code": "€", "Branch Name": "ANIMATION AND GRAPHICS"}, {"Branch Code": "AP", "Branch Name": "APPAREL TECHNOLOGY (SS)"}, {"Branch Code": "AR", "Branch Name": "ARCHITECTURE"}, {"Branch Code": "BA", "Branch Name": "ARCHITECTURE (SS)"}, {"Branch Code": "AD", "Branch Name": "ARTIFICIAL INTELLIGENCE AND DATA SCIENCE"}, {"Branch Code": "AT", "Branch Name": "ARTIFICIAL INTELLIGENCE AND DATA SCIENCE (SS)"}, {"Branch Code": "AL", "Branch Name": "ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING"}, {"Branch Code": "AU", "Branch Name": "AUTOMOBILE ENGINEERING"}, {"Branch Code": "AS", "Branch Name": "AUTOMOBILE ENGINEERING (SS)"}, {"Branch Code": "BP", "Branch Name": "B.PLAN"}, {"Branch Code": "DA", "Branch Name": "BACHELOR OF DESIGN"}, {"Branch Code": "BM", "Branch Name": "BIO MEDICAL ENGINEERING"}, {"Branch Code": "BY", "Branch Name": "BIO MEDICAL ENGINEERING (SS)"}, {"Branch Code": "BT", "Branch Name": "BIO TECHNOLOGY"}, {"Branch Code": "BS", "Branch Name": "BIO TECHNOLOGY (SS)"}, {"Branch Code": "BC", "Branch Name": "BIO TECHNOLOGY AND BIO CHEMICAL ENGINEERING"}, {"Branch Code": "CR", "Branch Name": "CERAMIC TECHNOLOGY (SS)"}, {"Branch Code": "CH", "Branch Name": "CHEMICAL ENGINEERING"}, {"Branch Code": "CL", "Branch Name": "CHEMICAL ENGINEERING (SS)"}, {"Branch Code": "CC", "Branch Name": "CHEMICAL AND ELECTRO CHEMICAL ENGINEERING (SS)"}, {"Branch Code": "CX", "Branch Name": "CHEMICAL AND ELECTROCHEMICAL ENGINEERING"}, {"Branch Code": "CE", "Branch Name": "CIVIL ENGINEERING"}, {"Branch Code": "CN", "Branch Name": "CIVIL ENGINEERING (SS)"}, {"Branch Code": "CZ", "Branch Name": "CIVIL AND STRUCTUTURAL ENGINEERING"}, {"Branch Code": "XC", "Branch Name": "CIVIL ENGINEERING (TAMIL MEDIUM)"}, {"Branch Code": "CK", "Branch Name": "CIVIL ENGINEERING (ENVIRONMENTAL ENGINEERING)"}, {"Branch Code": "CO", "Branch Name": "COMPUTER AND COMMUNICATION ENGINEERING"}, {"Branch Code": "CW", "Branch Name": "COMPUTER SCIENCE AND BUSINESS SYSTEM (SS)"}, {"Branch Code": "CB", "Branch Name": "COMPUTER SCIENCE AND BUSINESS SYSTEM"}, {"Branch Code": "CD", "Branch Name": "COMPUTER SCIENCE AND DESIGN"}, {"Branch Code": "CS", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING"}, {"Branch Code": "AM", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (AI & ML)"}, {"Branch Code": "CG", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (AI & ML) (SS)"}, {"Branch Code": "SC", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (CYBER SECURITY)"}, {"Branch Code": "CF", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE)"}, {"Branch Code": "SB", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (IOT + CYBER SECURITY + BLOCKCHAIN)"}, {"Branch Code": "CI", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS)"}, {"Branch Code": "CM", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (SS)"}, {"Branch Code": "XS", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (TAMIL)"}, {"Branch Code": "CA", "Branch Name": "COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE)"}, {"Branch Code": "TS", "Branch Name": "COMPUTER SCIENCE AND TECHNOLOGY"}, {"Branch Code": "CT", "Branch Name": "COMPUTER TECHNOLOGY"}, {"Branch Code": "CY", "Branch Name": "CYBER SECURITY"}, {"Branch Code": "EF", "Branch Name": "ELECTRICAL AND COMPUTER ENGINEERING"}, {"Branch Code": "ES", "Branch Name": "ELECTRICAL AND ELECTRONICS (SANDWICH) (SS)"}, {"Branch Code": "EE", "Branch Name": "ELECTRICAL AND ELECTRONICS ENGINEERING"}, {"Branch Code": "EY", "Branch Name": "ELECTRICAL AND ELECTRONICS ENGINEERING (SS)"}, {"Branch Code": "ET", "Branch Name": "ELECTRONICS AND TELECOMMUNICATION ENGINEERING"}, {"Branch Code": "EA", "Branch Name": "ELECTRONICS AND COMMUNICATION (ADVANCED COMMUNICATION TECHNOLOGY)"}, {"Branch Code": "EC", "Branch Name": "ELECTRONICS AND COMMUNICATION ENGINEERING"}, {"Branch Code": "EM", "Branch Name": "ELECTRONICS AND COMMUNICATION ENGINEERING (SS)"}, {"Branch Code": "EX", "Branch Name": "ELECTRONICS AND COMPUTER ENGINEERING"}, {"Branch Code": "EI", "Branch Name": "ELECTRONICS AND INSTRUMENTATION ENGINEERING"}, {"Branch Code": "EV", "Branch Name": "ELECTRONICS ENGINEERING (VLSI DESIGN AND TECHNOLOGY)"}, {"Branch Code": "EL", "Branch Name": "ELECTRONICS ENGINEERING (VLSI DESIGN AND TECHNOLOGY)"}, {"Branch Code": "IX", "Branch Name": "ELECTRONICS INSTRUMENTATION AND CONTROL ENGINEERING"}, {"Branch Code": "EN", "Branch Name": "ENVIRONMENTAL ENGINEERING"}, {"Branch Code": "FT", "Branch Name": "FASHION TECHNOLOGY"}, {"Branch Code": "FY", "Branch Name": "FASHION TECHNOLOGY (SS)"}, {"Branch Code": "FD", "Branch Name": "FOOD TECHNOLOGY"}, {"Branch Code": "FS", "Branch Name": "FOOD TECHNOLOGY (SS)"}, {"Branch Code": "GI", "Branch Name": "GEO INFORMATICS"}, {"Branch Code": "HT", "Branch Name": "HANDLOOM AND TEXTILE TECHNOLOGY"}, {"Branch Code": "IB", "Branch Name": "INDUSTRIAL BIO TECHNOLOGY"}, {"Branch Code": "IS", "Branch Name": "INDUSTRIAL BIO TECHNOLOGY (SS)"}, {"Branch Code": "IE", "Branch Name": "INDUSTRIAL ENGINEERING"}, {"Branch Code": "IN", "Branch Name": "INDUSTRIAL ENGINEERING AND MANAGEMENT"}, {"Branch Code": "SE", "Branch Name": "INFORMATION SCIENCE AND ENGINEERING"}, {"Branch Code": "IT", "Branch Name": "INFORMATION TECHNOLOGY"}, {"Branch Code": "IM", "Branch Name": "INFORMATION TECHNOLOGY (SS)"}, {"Branch Code": "IC", "Branch Name": "INSTRUMENTATION AND CONTROL ENGINEERING"}, {"Branch Code": "IY", "Branch Name": "INSTRUMENTATION AND CONTROL ENGINEERING (SS)"}, {"Branch Code": "ID", "Branch Name": "INTERIOR DESIGN (SS)"}, {"Branch Code": "LE", "Branch Name": "LEATHER TECHNOLOGY"}, {"Branch Code": "CJ", "Branch Name": "M.TECH COMPUTER SCIENCE AND ENGINEERING (INTEGRATED 5 YEARS)"}, {"Branch Code": "MN", "Branch Name": "MANUFACTURING ENGINEERING"}, {"Branch Code": "MR", "Branch Name": "MARINE ENGINEERING"}, {"Branch Code": "MA", "Branch Name": "MATERIAL SCIENCE AND ENGINEERING (SS)"}, {"Branch Code": "MU", "Branch Name": "MECHANICAL AND AUTOMATION ENGINEERING"}, {"Branch Code": "MO", "Branch Name": "MECHANICAL AND MECHATRONICS ENGINEERING (ADDITIVE MANUFACTURING)"}, {"Branch Code": "MJ", "Branch Name": "MECHANICAL AND SMART MANUFACTURING"}, {"Branch Code": "ME", "Branch Name": "MECHANICAL ENGINEERING"}, {"Branch Code": "MM", "Branch Name": "MECHANICAL ENGINEERING (MANUFACTURING)"}, {"Branch Code": "MH", "Branch Name": "MECHANICAL ENGINEERING (SANDWICH)"}, {"Branch Code": "MS", "Branch Name": "MECHANICAL ENGINEERING (SANDWICH) (SS)"}, {"Branch Code": "MF", "Branch Name": "MECHANICAL ENGINEERING (SS)"}, {"Branch Code": "XM", "Branch Name": "MECHANICAL ENGINEERING (TAMIL MEDIUM)"}, {"Branch Code": "MB", "Branch Name": "MECHANICAL ENGINEERING (AUTOMOBILE)"}, {"Branch Code": "MC", "Branch Name": "MECHATRONICS"}, {"Branch Code": "MG", "Branch Name": "MECHATRONICS (SS)"}, {"Branch Code": "MZ", "Branch Name": "MECHATRONICS ENGINEERING"}, {"Branch Code": "MD", "Branch Name": "MEDICAL ELECTRONICS ENGINEERING"}, {"Branch Code": "MT", "Branch Name": "METALLURGICAL ENGINEERING"}, {"Branch Code": "MY", "Branch Name": "METALLURGICAL ENGINEERING (SS)"}, {"Branch Code": "MI", "Branch Name": "MINING ENGINEERING"}, {"Branch Code": "PD", "Branch Name": "PETRO CHEMICAL ENGINEERING"}, {"Branch Code": "PC", "Branch Name": "PETRO CHEMICAL TECHNOLOGY"}, {"Branch Code": "PE", "Branch Name": "PETROLEUM ENGINEERING"}, {"Branch Code": "PP", "Branch Name": "PETROLEUM ENGINEERING AND TECHNOLOGY (SS)"}, {"Branch Code": "PH", "Branch Name": "PHARMACEUTICAL TECHNOLOGY"}, {"Branch Code": "PM", "Branch Name": "PHARMACEUTICAL TECHNOLOGY (SS)"}, {"Branch Code": "PA", "Branch Name": "PLASTIC TECHNOLOGY"}, {"Branch Code": "PT", "Branch Name": "PRINTING AND PACKING TECHNOLOGY"}, {"Branch Code": "PR", "Branch Name": "PRODUCTION ENGINEERING"}, {"Branch Code": "PS", "Branch Name": "PRODUCTION ENGINEERING (SANDWICH) (SS)"}, {"Branch Code": "PN", "Branch Name": "PRODUCTION ENGINEERING (SS)"}, {"Branch Code": "RI", "Branch Name": "ROBOTICS AND ARTIFICIAL INTELLIGENCE"}, {"Branch Code": "RM", "Branch Name": "ROBOTICS AND AUTOMATION"}, {"Branch Code": "RA", "Branch Name": "ROBOTICS AND AUTOMATION (SS)"}, {"Branch Code": "RP", "Branch Name": "RUBBER AND PLASTIC TECHNOLOGY"}, {"Branch Code": "SF", "Branch Name": "SAFETY AND FIRE ENGINEERING"}, {"Branch Code": "TC", "Branch Name": "TEXTILE CHEMISTRY"}, {"Branch Code": "TX", "Branch Name": "TEXTILE TECHNOLOGY"}, {"Branch Code": "TT", "Branch Name": "TEXTILE TECHNOLOGY (SS)"}, ]

    df_branch = pd.DataFrame(branch_data)

    colb1, colb2 = st.columns([2, 1])
    with colb1:
        branch_search = st.text_input("🔍 Search Branch", placeholder="ex: CS, AI, Civil, Mechanical...", key="branch_search")
    with colb2:
        st.write("")
        st.write(f"✅ Total Branches: **{len(df_branch)}**")

    df_branch_show = df_branch.copy()
    if branch_search.strip():
        df_branch_show = df_branch_show[
            df_branch_show["Branch Code"].str.contains(branch_search, case=False, na=False)
            | df_branch_show["Branch Name"].str.contains(branch_search, case=False, na=False)
        ]

    st.dataframe(df_branch_show, use_container_width=True, height=500)

    csv = df_branch.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Download Branch List (CSV)",
        data=csv,
        file_name="TNEA_Branch_List.csv",
        mime="text/csv"
    )

# =================================================
# ✅ PAGE 4: COLLEGE LIST (FREE)
# =================================================
elif selected == "TNEA College List (PDF)":

    import pdfplumber
    import re

    st.markdown("## 🏫 TNEA College List (Auto from PDF)")
    st.caption("✅ Extracted from official TNEA College PDF (S.No, College Code, College Name, Website, District)")

    # ✅ Your 2 PDF files (keep in same folder as app.py)
    PDF_FILES = [
        "TNEA_2025_College_full_list_1.pdf",
        "TNEA_2025_College_full_list_2.pdf"
    ]

    @st.cache_data(ttl=3600)
    def extract_college_list_from_multiple_pdfs(pdf_files):
        colleges = []
        s_no = 0

        for pdf_path in pdf_files:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue

                    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
                    if len(lines) == 0:
                        continue

                    # ✅ First line (College Code + College Name)
                    first_line = lines[0]

                    match = re.match(r"^(\d{1,4})\s+(.*)$", first_line)
                    if match:
                        college_code = match.group(1).strip()
                        college_name = match.group(2).strip()
                    else:
                        if re.match(r"^\d{1,4}$", first_line) and len(lines) > 1:
                            college_code = first_line.strip()
                            college_name = lines[1].strip()
                        else:
                            continue

                    # ✅ Extract District
                    district = ""
                    for ln in lines:
                        if ln.upper().startswith("DISTRICT"):
                            district = ln.replace("District", "").replace("DISTRICT", "").strip()
                            break

                    # ✅ Extract Website
                    website = ""
                    for ln in lines:
                        if ln.lower().startswith("website"):
                            website = ln.replace("Website", "").replace("WEBSITE", "").strip()
                            break

                    # ✅ Clean Website
                    website = website.replace(" ", "")
                    if website and not website.startswith("http"):
                        website = "https://" + website

                    s_no += 1
                    colleges.append({
                        "S.No": s_no,
                        "College Code": college_code,
                        "College Name": college_name,
                        "District": district.title() if district else "",
                        "Website": website
                    })

        df = pd.DataFrame(colleges)

        # ✅ Remove duplicates (sometimes PDF has repeated pages)
        df = df.drop_duplicates(subset=["College Code", "College Name"], keep="first")

        # ✅ Re-number S.No properly after removing duplicates
        df = df.reset_index(drop=True)
        df["S.No"] = df.index + 1

        return df

    # ✅ Load dataframe
    try:
        with st.spinner("📄 Reading college list from PDF files... Please wait"):
            df_college_pdf = extract_college_list_from_multiple_pdfs(PDF_FILES)

        st.success(f"✅ Total Colleges Extracted: {len(df_college_pdf)}")

    except Exception as e:
        st.error(f"❌ PDF Load Error: {e}")
        st.info("✅ Make sure BOTH PDF files exist in your app folder:")
        st.code("\n".join(PDF_FILES))
        st.stop()

    # ✅ Search + District filter
    col1, col2 = st.columns([2, 1])

    with col1:
        search_text = st.text_input(
            "🔍 Search College Name or College Code",
            placeholder="Ex: 1013, Guindy, MIT...",
            key="college_pdf_search"
        )

    with col2:
        district_list = ["All"] + sorted([
            d for d in df_college_pdf["District"].dropna().unique().tolist()
            if str(d).strip() != ""
        ])
        selected_district = st.selectbox(
            "📍 Filter by District",
            district_list,
            key="college_pdf_district"
        )

    df_show = df_college_pdf.copy()

    if selected_district != "All":
        df_show = df_show[df_show["District"] == selected_district]

    if search_text.strip():
        df_show = df_show[
            df_show["College Name"].str.contains(search_text, case=False, na=False)
            | df_show["College Code"].astype(str).str.contains(search_text, case=False, na=False)
        ]

    st.write(f"✅ Colleges Found: **{len(df_show)}**")

    st.dataframe(df_show, use_container_width=True, height=500)

  # ✅ Optional: Open website button list
st.markdown("---")
st.markdown("### 🌐 Open College Websites")

for _, row in df_show.head(50).iterrows():  # show only first 50
    st.markdown(f"**{row['College Code']} - {row['College Name']}** ({row['District']})")

    website = row.get("Website", "")

    # ✅ Fix NaN / None / invalid values
    if pd.notna(website) and str(website).strip() != "":
        website = str(website).strip()

        # ✅ Add https if missing
        if not website.startswith("http"):
            website = "https://" + website

        st.link_button(
            "🌐 Open Website",
            website,
            key=f"open_{row['College Code']}_{row['S.No']}"
        )
    else:
        st.info("Website not available in PDF")

    st.markdown("---")

# =================================================
# ✅ PAGE 5: CHOICE LIST (PREMIUM ONLY)
# =================================================
elif selected == "2024-TNEA  Cut off and rank details":

    # ✅ Premium restriction
    if not is_premium:
        st.warning("🔒 Premium only feature. Click 💳 Go Premium button on top.")
        st.stop()

    import time

    # ✅ Performance-optimized loader with caching (10 minutes)
    @st.cache_data(ttl=600)
    def load_excel_file_from_url(url):
        response = requests.get(url)
        df_loaded = pd.read_excel(io.BytesIO(response.content))
        return df_loaded

    excel_url = "https://docs.google.com/spreadsheets/d/1rASGgYC9RZA0vgmtuFYRG0QO3DOGH_jW/export?format=xlsx"

    with st.spinner("📥 Loading TNEA cutoff data..."):
        df = load_excel_file_from_url(excel_url)

    for col in df.columns:
        if col.endswith("_C") or col.endswith("_GR"):
            df[col] = pd.to_numeric(df[col], errors="coerce")

    st.image("https://drive.google.com/thumbnail?id=1FPfkRH3BC1BeQRtQVpZDH3P3ilTSMYNA", width=100)
    st.title("📊 TNEA 2025 Cutoff & Rank Finder")
    st.markdown(f"🆔 **Accessed by: {st.session_state.mobile}**")

    df['College_Option'] = df['CL'].astype(str) + " - " + df['College']
    college_options = sorted(df['College_Option'].unique().tolist())
    selected_college = st.selectbox("🏛️ Select College", options=["All"] + college_options)

    st.subheader("🎯 Filter by Community, Department, Zone")
    if selected_college == "All":
        community = st.selectbox(
            "Select Community",
            options=["All", "OC", "BC", "BCM", "MBC", "SC", "SCA", "ST"],
            key="main_community"
        )
        department = st.selectbox(
            "Select Department (Br)",
            options=["All"] + sorted(df['Br'].dropna().unique().tolist())
        )
        zone = st.selectbox(
            "Select Zone",
            options=["All"] + sorted(df['zone'].dropna().unique().tolist())
        )

    st.subheader("📌 Compare Up to 5 Colleges")
    compare_colleges = st.multiselect(
        "Select colleges to compare",
        options=college_options,
        max_selections=5
    )

    if compare_colleges:
        st.markdown("### 🎯 Filter Inside Compared Colleges")
        comp_dept = st.selectbox(
            "Department",
            options=["All"] + sorted(df['Br'].dropna().unique().tolist()),
            key="compare_department"
        )
        comp_comm = st.selectbox(
            "Community",
            options=["All", "OC", "BC", "BCM", "MBC", "SC", "SCA", "ST"],
            key="compare_community"
        )

        compare_cls = [c.split(" - ")[0].strip() for c in compare_colleges]
        compare_df = df[df['CL'].astype(str).isin(compare_cls)]

        if comp_dept != "All":
            compare_df = compare_df[compare_df['Br'] == comp_dept]

        color_palette = ['#f7c6c7', '#c6e2ff', '#d5f5e3', '#fff5ba', '#e0ccff']
        college_color_map = {cl: color_palette[i] for i, cl in enumerate(compare_cls)}

        def highlight_college(row):
            cl = str(row['CL'])
            bg_color = college_color_map.get(cl, '#ffffff')
            return [f'background-color: {bg_color}; color: black;' for _ in row]

        compare_cols = ['CL', 'College', 'Br', 'zone']
        if comp_comm != "All":
            compare_cols += [f"{comp_comm}_C", f"{comp_comm}_GR"]
        else:
            compare_cols += [col for col in df.columns if col.endswith("_C") or col.endswith("_GR")]

        format_dict = {
            col: '{:.2f}' if '_C' in col else '{:.0f}'
            for col in compare_cols
            if '_C' in col or '_GR' in col
        }

        st.markdown("### 🟨 College Comparison Table")
        st.dataframe(
            compare_df[compare_cols]
            .style
            .apply(highlight_college, axis=1)
            .format(format_dict)
            .hide(axis='index'),
            height=450
        )

    # --- MAIN FILTERED DATA ---
    show_data = False
    filtered_df = df.copy()

    if selected_college != "All":
        show_data = True
        selected_cl = selected_college.split(" - ")[0].strip()
        filtered_df = filtered_df[filtered_df['CL'].astype(str) == selected_cl]
    else:
        if 'zone' in locals() and zone != "All":
            filtered_df = filtered_df[filtered_df['zone'] == zone]
            show_data = True
        if 'department' in locals() and department != "All":
            filtered_df = filtered_df[filtered_df['Br'] == department]
            show_data = True

    if selected_college == "All" and 'community' in locals() and community != "All":
        cols_to_show = ['CL', 'College', 'Br', f'{community}_C', f'{community}_GR', 'zone']
    else:
        cols_to_show = ['CL', 'College', 'Br', 'zone'] + [
            col for col in df.columns if col.endswith("_C") or col.endswith("_GR")
        ]

    format_dict = {
        col: '{:.2f}' if '_C' in col else '{:.0f}'
        for col in cols_to_show
        if '_C' in col or '_GR' in col
    }

    st.markdown("### 🔎 Filtered Results")
    if show_data:
        st.dataframe(
            filtered_df[cols_to_show]
            .style
            .format(format_dict)
            .hide(axis='index'),
            height=600
        )
    else:
        st.info("Please apply filters to see the results.")

# =================================================
# ✅ PAGE 6: TNEA VACANCY SEAT MATRIX (PREMIUM ONLY)
# =================================================
elif selected == "2025-TNEA Vacancy Seat Matrix":

    # ✅ Premium restriction
    if not is_premium:
        st.warning("🔒 Premium only feature. Click 💳 Go Premium button (top right).")
        st.stop()

    st.title("📊 2025-TNEA Vacancy Seat Matrix")
    st.success("✅ Premium Feature Enabled ✅")

    import plotly.express as px
    from openpyxl import load_workbook
    import requests
    import io

    # ✅ All Google Sheet URLs
    excel_urls = {
        "Round 1": "https://docs.google.com/spreadsheets/d/17otzGFO0AhKzx5ChSUhW18HnqA8Ed2sY/export?format=xlsx",
        "Round 2": "https://docs.google.com/spreadsheets/d/1H1pLjbsvaOl1UMBAJbtfWz1B-KZQ24iB/export?format=xlsx",
        "Round 3": "https://docs.google.com/spreadsheets/d/1VPfuYg6cNtm_x4gnGkEndssR8CqfCDJT/export?format=xlsx",
        "Supplymentry_Counselling": "https://docs.google.com/spreadsheets/d/1NEf4pHVjO1m0Lz1g3Lua3emXT39opYbV/export?format=xlsx",
        "After_Supplymentry_Counselling": "https://docs.google.com/spreadsheets/d/1qnsPSyCd-myYnsOcF8vv_H-gPwn6psL8/export?format=xlsx",
        "After_SCA_to_SC": "https://docs.google.com/spreadsheets/d/1yqCtM98OC0GJSaAt-gbcVy8eTcWKc5pD/export?format=xlsx",
    }

    @st.cache_data(ttl=600)
    def load_all_rounds():
        all_data = {}
        for round_name, excel_url in excel_urls.items():
            try:
                response = requests.get(excel_url)
                excel_file = io.BytesIO(response.content)
                wb = load_workbook(excel_file, data_only=True)

                data_dict = {}
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    data = list(ws.values)
                    if not data:
                        continue
                    header = data[0]
                    rows = data[1:]
                    df_temp = pd.DataFrame(rows, columns=header)
                    data_dict[sheet] = df_temp

                all_data[round_name] = data_dict

            except Exception as e:
                st.error(f"⚠️ Error loading {round_name}: {e}")

        return all_data

    all_rounds_data = load_all_rounds()

    community_cols = ['OC', 'BC', 'BCM', 'MBC', 'SC', 'SCA', 'ST']
    required_id_vars = ['College Name', 'College Code', 'Branch Code', 'Branch Name']

    # ----------------------------- CATEGORY 1 -----------------------------
    st.markdown("## 📂 Select Round, Branch and Community")
    col_cat1_0, col_cat1_1, col_cat1_2, col_cat1_3 = st.columns(4)

    with col_cat1_0:
        selected_round_1 = st.selectbox(
            "📂 Select Counselling Round",
            list(all_rounds_data.keys()),
            key="cat1_round"
        )

    with col_cat1_1:
        selected_sheet_1 = st.selectbox(
            "📂 Select Vacancy - Category",
            list(all_rounds_data[selected_round_1].keys()),
            key="cat1_sheet"
        )
        df1 = all_rounds_data[selected_round_1][selected_sheet_1]

    if df1.empty:
        st.error("❌ No data found in the selected sheet.")
        st.stop()

    # ✅ Clean dataframe
    df1.columns = [str(col).strip().upper().replace("  ", " ").replace("\n", " ") for col in df1.columns]

    rename_map = {}
    for col in df1.columns:
        if "COLLEGE CODE" in col:
            rename_map[col] = 'College Code'
        elif "COLLEGE NAME" in col:
            rename_map[col] = 'College Name'
        elif "BRANCH CODE" in col:
            rename_map[col] = 'Branch Code'
        elif "BRANCH NAME" in col:
            rename_map[col] = 'Branch Name'

    df1.rename(columns=rename_map, inplace=True)

    df1 = df1[[col for col in df1.columns if col in required_id_vars + community_cols]]
    df1[community_cols] = df1[community_cols].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)

    # ✅ Total Seats (Round-wise)
    total_round_seats = df1[community_cols].sum().sum()

    # ✅ Total per Branch
    branch_totals = df1.groupby("Branch Code")[community_cols].sum()
    branch_totals["Total Vacant"] = branch_totals.sum(axis=1)

    st.metric(label=f"🎯 Total Seats in {selected_round_1}", value=f"{total_round_seats:,}")

    with st.expander("📊 Branch-wise Vacancy Summary"):
        st.dataframe(branch_totals.reset_index(), use_container_width=True)

    # ----------------- Filters -----------------
    with col_cat1_2:
        branch_codes = sorted(df1['Branch Code'].dropna().unique())
        selected_branch_1 = st.selectbox("🔍 Select Branch Code", ['All'] + branch_codes)

    with col_cat1_3:
        selected_community_1 = st.selectbox("🧑‍🤝‍🧑 Select Community", ['All'] + community_cols)

    if selected_branch_1 == 'All':
        branch_df = df1.copy()
    else:
        branch_df = df1[df1['Branch Code'] == selected_branch_1].copy()

    if selected_community_1 == 'All':
        branch_df['Total Seats (All Communities)'] = branch_df[community_cols].sum(axis=1)
        branch_df = branch_df[[*required_id_vars, 'Total Seats (All Communities)']]
    else:
        branch_df = branch_df[[*required_id_vars, selected_community_1]]
        branch_df = branch_df.rename(columns={selected_community_1: 'Selected Community Seats'})
        branch_df.insert(4, 'Selected Community', selected_community_1)

    if not branch_df.empty:
        title_text = f"📘 Round: {selected_round_1} | Branch: {selected_branch_1} | Community: {selected_community_1}"
        st.header(title_text)
        st.dataframe(branch_df, use_container_width=True)

        if selected_branch_1 != 'All':
            bar1 = df1[df1['Branch Code'] == selected_branch_1]
            community_summary = bar1[community_cols].sum().reset_index()
            community_summary.columns = ['Community', 'Seats']

            total_branch_seats = community_summary['Seats'].sum()

            chart_title = (
                f"{selected_round_1} - {selected_sheet_1} - {selected_branch_1} "
                f"- Total Seats Across Communities (Total = {total_branch_seats:,})"
            )

            fig1 = px.bar(
                community_summary,
                x='Community',
                y='Seats',
                color='Community',
                text='Seats',
                title=chart_title,
                labels={'Community': 'Community Category', 'Seats': 'Number of Seats'},
                height=450
            )
            fig1.update_layout(xaxis_title="Community", yaxis_title="Number of Seats")
            fig1.update_traces(textposition='outside')
            st.plotly_chart(fig1, use_container_width=True)

    # ----------------------------- CATEGORY 2 -----------------------------
    st.markdown("---")
    st.markdown("## 🏧 Select Round, College and Branch")
    col_cat2_0, col_cat2_1, col_cat2_2, col_cat2_3 = st.columns(4)

    with col_cat2_0:
        selected_round_2 = st.selectbox(
            "📂 Select Counselling Round",
            list(all_rounds_data.keys()),
            key="cat2_round"
        )

    with col_cat2_1:
        selected_sheet_2 = st.selectbox(
            "📂 Select Vacancy - Category",
            list(all_rounds_data[selected_round_2].keys()),
            key="cat2_sheet"
        )
        df2 = all_rounds_data[selected_round_2][selected_sheet_2]

    if df2.empty:
        st.error("❌ No data found in the selected sheet.")
        st.stop()

    df2.columns = [str(col).strip().upper().replace("  ", " ").replace("\n", " ") for col in df2.columns]
    df2.rename(columns=rename_map, inplace=True)

    df2 = df2[[col for col in df2.columns if col in required_id_vars + community_cols]]
    df2[community_cols] = df2[community_cols].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)

    df2['College Combined'] = df2['College Code'].astype(str) + ' - ' + df2['College Name']

    unique_colleges = sorted(df2['College Combined'].dropna().unique())

    with col_cat2_2:
        selected_college_combined = st.selectbox("🏧 Select College (Code - Name)", ['All'] + unique_colleges)

    with col_cat2_3:
        branch_codes_2 = sorted(df2['Branch Code'].dropna().unique())
        selected_branch_code = st.selectbox("🔍 Filter by Branch Code (Optional)", ['All'] + branch_codes_2)

    selected_community_2 = st.selectbox("🧑‍🤝‍🧑 Select Community", ['All'] + community_cols, key="cat2_community")

    college_df = df2.copy()

    if selected_college_combined != "All":
        selected_code, selected_name = selected_college_combined.split(" - ", 1)
        college_df = college_df[
            (college_df['College Code'].astype(str) == selected_code.strip()) &
            (college_df['College Name'].str.strip() == selected_name.strip())
        ]

    if selected_branch_code != "All":
        college_df = college_df[college_df['Branch Code'] == selected_branch_code]

    # Chart 1: All Community Seats per Branch
    fig2_data_all = college_df.copy()
    fig2_data_all['Total Seats (All Communities)'] = fig2_data_all[community_cols].sum(axis=1)

    fig_all = px.bar(
        fig2_data_all,
        x='Branch Code',
        y='Total Seats (All Communities)',
        color='Branch Code',
        text='Total Seats (All Communities)',
        title=f"{selected_round_2} - {selected_college_combined} - Total Seats per Branch (All Communities)",
        labels={'Branch Code': 'Branch', 'Total Seats (All Communities)': 'Number of Seats'},
        height=450
    )
    fig_all.update_layout(xaxis_title="Branch", yaxis_title="Number of Seats")
    fig_all.update_traces(textposition='outside')
    st.plotly_chart(fig_all, use_container_width=True)

    # Chart 2: Selected Community Seats per Branch
    if selected_community_2 != 'All':
        college_df2 = college_df[[*required_id_vars, selected_community_2]].copy()
        college_df2 = college_df2.rename(columns={selected_community_2: 'Selected Community Seats'})
        college_df2.insert(4, 'Selected Community', selected_community_2)

        fig2 = px.bar(
            college_df2,
            x='Branch Code',
            y='Selected Community Seats',
            color='Branch Code',
            text='Selected Community Seats',
            title=f"{selected_round_2} - {selected_college_combined} - {selected_community_2} Seats per Branch",
            labels={'Branch Code': 'Branch', 'Selected Community Seats': 'Number of Seats'},
            height=450
        )
        fig2.update_layout(xaxis_title="Branch", yaxis_title="Number of Seats")
        fig2.update_traces(textposition='outside')
        st.plotly_chart(fig2, use_container_width=True)

    if college_df.empty:
        st.warning("⚠️ No data found for the selected college or branch.")


















